#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Wandelt eine ausgefüllte Excel-Erhebungs-Vorlage in YAML-Datenelement-ENTWÜRFE um.

Die erzeugten YAMLs sind bewusst nur ein Ausgangspunkt (publication_status: AuthorDraft):
Codings, Standard-Mappings und die endgültige semantische Definition bleiben Aufgabe des
MI-Teams. Es wird NICHTS committet und der Bestand unter `elements/` wird NICHT überschrieben —
die Entwürfe landen in einem Staging-Verzeichnis (Default: `elements/_incoming/`, gitignored).

Aufruf:
  pip install openpyxl
  python scripts/import-elicitation-workbook.py ausgefuellt.xlsx
  python scripts/import-elicitation-workbook.py ausgefuellt.xlsx --out /tmp/drafts --validate

Danach Dubletten prüfen:  python scripts/check-duplicates.py --candidates <out-dir>
"""
from __future__ import annotations

import argparse
import re
import subprocess
import sys
import unicodedata
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

import yaml
from openpyxl import load_workbook

import _elicitation as E

_UMLAUT = {"ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss", "Ä": "Ae", "Ö": "Oe", "Ü": "Ue"}


def camel_name(label: str) -> str:
    """Erzeugt einen schema-konformen camelCase-`name` aus einer deutschen Bezeichnung."""
    for k, v in _UMLAUT.items():
        label = label.replace(k, v)
    label = unicodedata.normalize("NFKD", label).encode("ascii", "ignore").decode()
    words = [w for w in re.split(r"[^A-Za-z0-9]+", label) if w]
    if not words:
        return "neuesDatenelement"
    head, *tail = words
    name = head[:1].lower() + head[1:] + "".join(w[:1].upper() + w[1:] for w in tail)
    if not re.match(r"^[a-z]", name):
        name = "x" + name
    return re.sub(r"[^A-Za-z0-9]", "", name)


def _rows(ws, columns) -> list[dict]:
    keys = [c[0] for c in columns]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        out.append({keys[i]: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                    for i in range(len(keys))})
    return out


def build_flows(flow_rows: list[dict], local_id: str) -> list[dict]:
    flows = []
    for fr in flow_rows:
        if fr.get("local_id") != local_id:
            continue
        system = E.map_value(fr.get("system"), E.SYSTEM_LAYMEN, default="other")
        usage = E.map_value(fr.get("usage_type"), E.USAGE_LAYMEN, default="other")
        flow = {"system": system, "usage_type": usage}
        if fr.get("system_label"):
            flow["system_label"] = fr["system_label"]
        if fr.get("actor_role"):
            flow["actor_role"] = fr["actor_role"]
        sector = E.map_value(fr.get("sector"), E.SECTOR_LAYMEN, default="")
        if sector:
            flow["sector"] = sector
        mand = E.map_value(fr.get("mandatory"), E.JA_NEIN, default=None)
        if isinstance(mand, bool):
            flow["mandatory"] = mand
        if fr.get("notes"):
            flow["notes"] = fr["notes"]
        flows.append(flow)
    return flows


def build_codings(code_rows: list[dict], local_id: str) -> list[dict]:
    """value_set.codings[] aus dem Blatt 'Codes' (System + Code Pflicht)."""
    codings = []
    for cr in code_rows:
        if cr.get("local_id") != local_id:
            continue
        system, code = cr.get("terminology_system", "").strip(), cr.get("code", "").strip()
        if not (system and code):
            continue
        c = {"system": system, "code": code}
        if cr.get("display"):
            c["display"] = cr["display"]
        codings.append(c)
    return codings


def to_element(er: dict, flow_rows: list[dict], code_rows: list[dict], today: str) -> dict:
    phase = E.map_value(er.get("phase"), E.PHASE_LAYMEN, default="")
    if phase == "other" or not phase:
        phase = re.sub(r"[^a-z]", "", (er.get("phase_other") or "newphase").lower()) or "newphase"
    name = camel_name(er.get("label_de", ""))
    datatype = E.map_value(er.get("datatype"), E.DATATYPE_LAYMEN, default="string")
    cardinality = E.map_value(er.get("cardinality"), E.CARDINALITY_LAYMEN, default="0..1")

    doc: dict = {
        "id": f"mihub.lung.{phase}.{name}.v0.1",
        "uuid": str(uuid.uuid4()),
        "name": name,
        "label_de": er.get("label_de", ""),
        "definition_de": er.get("definition_de", ""),
        "phase": phase,
        "datatype": datatype,
        "cardinality": cardinality,
    }
    if er.get("tumor_entity"):
        ents = [e.strip() for e in er["tumor_entity"].replace(";", ",").split(",") if e.strip()]
        doc["tumor_entity"] = ents or ["lung_cancer_any"]
    if er.get("unit"):
        doc["unit"] = er["unit"]

    # value_set: Binding + Status (Roh-Enums aus strikten Dropdowns) + Codings (Blatt 'Codes')
    codings = build_codings(code_rows, er.get("local_id", ""))
    vs: dict = {}
    if er.get("binding_strength"):
        vs["binding_strength"] = er["binding_strength"]
    if er.get("standard_binding_status"):
        vs["standard_binding_status"] = er["standard_binding_status"]
    if codings:
        vs["codings"] = codings
    if vs:
        doc["value_set"] = vs

    # standard_mappings: primäres Ziel (MI-Team ergänzt Profil/Pfad)
    if er.get("standard_primary"):
        doc["standard_mappings"] = [{"standard": er["standard_primary"]}]

    care: dict = {}
    for k_src, k_dst in (("trigger", "trigger"), ("frequency_pattern", "frequency_pattern"),
                         ("responsible_role", "responsible_role")):
        if er.get(k_src):
            care[k_dst] = er[k_src]
    flows = build_flows(flow_rows, er.get("local_id", ""))
    if flows:
        care["data_flows"] = flows
    if care:
        doc["care_process"] = care

    if er.get("guideline_source") or er.get("guideline_section"):
        ref = {"source": er.get("guideline_source") or "andere",
               "version": "TBD — vom MI-Team zu prüfen",
               "section": er.get("guideline_section", "")}
        if er.get("recommendation_grade"):
            ref["recommendation_grade"] = er["recommendation_grade"]
        if er.get("evidence_level"):
            ref["evidence_level"] = er["evidence_level"]
        doc["evidence"] = {"guideline_references": [ref]}

    issues = []
    if er.get("notes"):
        issues.append(er["notes"])
    issues.append("Aus Excel-Erhebung importiert — Codes/Mappings/Definition durch MI-Team zu finalisieren.")
    doc["issues"] = issues

    doc["iso13972_metadata"] = {
        "keywords": ["Lungenkarzinom", phase],
        "version_number": "0.1.0",
        "creation_date": today,
        "publication_status": "AuthorDraft",
        "publisher": "MiHUB Konsortium AP8",
        "language": ["de", "en"],
        "license": "CC-BY-4.0",
    }
    doc["provenance"] = {
        "created_by": er.get("contact") or "Klinik-Spur (Excel-Import)",
        "reviewers": [],
        "git_commit": "",
    }
    return doc


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook")
    ap.add_argument("--out", default=str(E.ROOT / "elements" / "_incoming"))
    ap.add_argument("--validate", action="store_true", help="scripts/validate.py über die Entwürfe laufen lassen")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, data_only=True)
    if "Datenelemente" not in wb.sheetnames:
        print("FEHLER: Blatt 'Datenelemente' nicht gefunden.", file=sys.stderr)
        return 2
    elem_rows = _rows(wb["Datenelemente"], E.ELEMENT_COLUMNS)
    flow_rows = _rows(wb["Datennutzung"], E.FLOW_COLUMNS) if "Datennutzung" in wb.sheetnames else []
    code_rows = _rows(wb["Codes"], E.CODE_COLUMNS) if "Codes" in wb.sheetnames else []
    today = datetime.now(timezone.utc).date().isoformat()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for er in elem_rows:
        if E.is_example(er.get("local_id")):
            print(f"  ignoriert (Beispielzeile): {er.get('local_id')}")
            continue
        if not (er.get("label_de") and er.get("definition_de")):
            print(f"  übersprungen (Pflichtfelder fehlen): {er.get('local_id') or er.get('label_de')}")
            continue
        doc = to_element(er, flow_rows, code_rows, today)
        fp = out_dir / f"{doc['phase']}__{doc['name']}.yaml"
        fp.write_text(yaml.safe_dump(doc, allow_unicode=True, sort_keys=False), encoding="utf-8")
        written.append(fp)
        print(f"  -> {fp.relative_to(E.ROOT) if E.ROOT in fp.parents else fp}  (id: {doc['id']})")

    print(f"\n{len(written)} Entwurf/Entwürfe in {out_dir}")
    print("Nächste Schritte: 1) Dubletten prüfen  python scripts/check-duplicates.py --candidates "
          f"{out_dir}\n                  2) MI-Team: Codes/Mappings ergänzen, dann nach elements/<phase>/ verschieben")

    if args.validate and written:
        print("\n--- Schema-Validierung der Entwürfe ---")
        subprocess.run([sys.executable, str(E.ROOT / "scripts" / "validate.py"), *map(str, written)])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
