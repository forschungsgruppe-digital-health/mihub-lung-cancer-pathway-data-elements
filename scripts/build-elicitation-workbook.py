#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Erzeugt die Excel-Erhebungs-Vorlage für die Klinik-Spur (ohne Git/YAML).

Schema-getrieben: alle Dropdown-Werte stammen aus den Enums in
`schemas/data-element.schema.json` bzw. den Listen in `scripts/_elicitation.py` — die Vorlage
kann daher nicht vom Schema abdriften. Selbsterklärend durch Anleitung, Spalten-Kommentare,
Pflicht-Markierung und eine Beispielzeile (wird beim Import ignoriert).

Aufruf:
  pip install openpyxl
  python scripts/build-elicitation-workbook.py                 # -> templates/datenelement-erhebung.xlsx
  python scripts/build-elicitation-workbook.py --out x.xlsx --rows 80

Ausgefüllte Mappen werden mit `scripts/import-elicitation-workbook.py` zu YAML-Entwürfen.
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import _elicitation as E

HEADER_FILL = PatternFill("solid", fgColor="1A5DAB")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MAND_FILL = PatternFill("solid", fgColor="FCE8E6")   # zarte Markierung für Pflichtspalten
EXAMPLE_FILL = PatternFill("solid", fgColor="EFEFEF")  # graue Beispielzeile
EXAMPLE_FONT = Font(italic=True, color="666666")
WRAP = Alignment(wrap_text=True, vertical="top")
TITLE_FONT = Font(bold=True, size=14)


def _style_header(ws, columns, help_map) -> None:
    for c, (key, _h, _m, _l) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        tip = help_map.get(key)
        if tip:
            cmt = Comment(tip, "MiHUB")
            cmt.width, cmt.height = 320, 170
            cell.comment = cmt
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 34


def _write_codelists(ws, lists: dict[str, list[str]]) -> dict[str, str]:
    """Schreibt jede Dropdown-Liste in eine Spalte; gibt {listname: range-formula} zurück."""
    ranges: dict[str, str] = {}
    for idx, (name, values) in enumerate(lists.items(), start=1):
        col = get_column_letter(idx)
        ws.cell(row=1, column=idx, value=name).font = Font(bold=True)
        for r, val in enumerate(values, start=2):
            ws.cell(row=r, column=idx, value=val)
        ranges[name] = f"=Codelisten!${col}$2:${col}${len(values) + 1}"
        ws.column_dimensions[col].width = max(18, min(52, max((len(v) for v in values), default=10) + 2))
    ws.sheet_state = "hidden"
    return ranges


def _add_table(wb, sheet_name, columns, data_rows, ranges, extra_list_validations=None):
    """Legt ein Tabellenblatt an (Kopf + Kommentare + Pflicht-Fills + Dropdowns über die Datenzeilen)."""
    ws = wb.create_sheet(sheet_name)
    for ci, (_key, header, mandatory, _list) in enumerate(columns, start=1):
        ws.cell(row=1, column=ci, value=header)
        wide = header.startswith(("Klinische_Definition", "Anmerkungen", "Hinweis"))
        ws.column_dimensions[get_column_letter(ci)].width = 42 if wide else 24
        if mandatory:
            for r in range(2, data_rows + 2):
                ws.cell(row=r, column=ci).fill = MAND_FILL
    _style_header(ws, columns, E.HELP)
    last = data_rows + 1
    for ci, (_key, _header, _mandatory, listname) in enumerate(columns, start=1):
        if not listname:
            continue
        formula = ranges.get(listname) or (extra_list_validations or {}).get(listname)
        if not formula:
            continue
        suggestion = listname in E.SUGGESTION_LISTS
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showErrorMessage=not suggestion)
        if not suggestion:
            dv.errorTitle = "Ungültige Eingabe"
            dv.error = "Bitte einen Wert aus der Dropdown-Liste wählen."
            dv.errorStyle = "stop"
        col = get_column_letter(ci)
        dv.add(f"{col}2:{col}{last}")
        ws.add_data_validation(dv)
    return ws


def _write_example_row(ws, columns, row_idx, example: dict) -> None:
    """Schreibt eine grau-kursive Beispielzeile + Hinweis-Kommentar auf der lokalen ID."""
    for ci, (key, _h, _m, _l) in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=ci, value=example.get(key, ""))
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.alignment = WRAP
    note = Comment(
        "BEISPIELZEILE — wird beim Import IGNORIERT (lokale ID beginnt mit 'BEISPIEL').\n"
        "Sie können sie löschen oder einfach stehen lassen. Eigene Einträge in die Zeilen "
        "darunter (DE-001, DE-002, …).", "MiHUB")
    note.width, note.height = 340, 130
    ws.cell(row=row_idx, column=1).comment = note


def _write_intro(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 118
    B = Font(bold=True, size=12)
    lines = [
        ("MiHUB Lungenkrebs — Datenelement-Erhebung (Klinik-Spur)", TITLE_FONT),
        ("", None),
        ("Mit dieser Mappe schlagen klinische Expert:innen neue Datenelemente für den "
         "Lungenkrebs-Patientenpfad vor — ohne GitHub, ohne YAML. Das MI-Team übernimmt danach "
         "die technische Umsetzung (Codes, FHIR-Mappings) und bittet Sie um inhaltliche Bestätigung.", None),
        ("", None),
        ("Die Blätter dieser Mappe", B),
        ("• »Datenelemente« — eine Zeile je Datenelement (Hauptblatt).", None),
        ("• »Datennutzung« — WO (System) · WER (Rolle) · WIE (Nutzung); mehrere Zeilen je Element.", None),
        ("• »Codes« (optional) — bekannte Codes je Element, inkl. Terminologie-System (Dropdown).", None),
        ("• »Codelisten« — ausgeblendet; speist nur die Dropdowns (bitte nicht bearbeiten).", None),
        ("", None),
        ("Legende", B),
        ("• Spalten mit * und rosa Hintergrund sind PFLICHT.", None),
        ("• Die erste, grau-kursive Zeile (lokale ID = BEISPIEL) ist ein BEISPIEL und wird beim "
         "Import IGNORIERT — gerne überschreiben oder löschen.", None),
        ("• Felder mit Dropdown: Pfeil rechts in der Zelle. Manche Listen sind verbindlich, "
         "andere nur Vorschläge (dann ist auch Freitext erlaubt).", None),
        ("• Erklärung je Spalte: Mauszeiger über die Kopfzeile (rotes Eck = Kommentar).", None),
        ("", None),
        ("So füllen Sie aus", B),
        ("1. »Datenelemente«: je Element eine Zeile; Pflichtfelder ausfüllen, Dropdowns nutzen.", None),
        ("2. Vergeben Sie je Zeile eine kurze »lokale ID« (z. B. DE-001) — sie verknüpft die "
         "Blätter »Datennutzung« und »Codes« mit dem Datenelement.", None),
        ("3. Optional »Datennutzung« und »Codes« je lokale ID befüllen.", None),
        ("4. Was Sie nicht wissen (z. B. SNOMED-Codes, Standard-Mapping), darf leer bleiben — "
         "das MI-Team recherchiert es.", None),
        ("5. Mappe an digital-health@tu-dresden.de senden oder an ein GitHub-Issue anhängen.", None),
        ("", None),
        ("Die ausgefüllte Beispielzeile »Rauchstatus« in jedem Blatt zeigt eine vollständige Erfassung.", Font(italic=True)),
        ("", None),
        ("Beiträge erfolgen unter CC BY 4.0. Bitte KEINE personenbezogenen / realen Patientendaten "
         "eintragen — nur die abstrakte Datenelement-Definition.", Font(italic=True)),
        ("Andere Wege: GitHub-Issue-Formular (online) oder YAML+Pull-Request (technisch) — "
         "siehe CONTRIBUTING.md. Kontakt: digital-health@tu-dresden.de", None),
    ]
    for r, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=r, column=2, value=text)
        cell.alignment = WRAP
        if font:
            cell.font = font


def build(out_path: Path, rows: int) -> Path:
    enums = E.load_schema_enums()
    lists = E.dropdown_lists(enums)

    wb = Workbook()
    _write_intro(wb.active)
    wb.active.title = "Anleitung"

    code_src = wb.create_sheet("Codelisten")  # zunächst leer; nach den Tabellen befüllt

    # --- Datenelemente: 1 Beispielzeile + `rows` echte Zeilen ---
    elem_rows = rows + 1
    elem_ws = _add_table(wb, "Datenelemente", E.ELEMENT_COLUMNS, elem_rows, {})
    _write_example_row(elem_ws, E.ELEMENT_COLUMNS, 2, E.EXAMPLE_ELEMENT)
    for i in range(rows):
        elem_ws.cell(row=3 + i, column=1, value=f"DE-{i + 1:03d}")

    ranges = _write_codelists(code_src, lists)
    id_range = {"element_ids": f"=Datenelemente!$A$2:$A${elem_rows + 1}"}

    # --- Datennutzung + Codes (local_id als Liste der Datenelement-IDs) ---
    flow_rows = rows * 3 + len(E.EXAMPLE_FLOWS)
    flow_ws = _add_table(wb, "Datennutzung", E.FLOW_COLUMNS, flow_rows, ranges, id_range)
    for i, ex in enumerate(E.EXAMPLE_FLOWS):
        _write_example_row(flow_ws, E.FLOW_COLUMNS, 2 + i, ex)

    code_rows = rows * 2 + len(E.EXAMPLE_CODES)
    codes_ws = _add_table(wb, "Codes", E.CODE_COLUMNS, code_rows, ranges, id_range)
    for i, ex in enumerate(E.EXAMPLE_CODES):
        _write_example_row(codes_ws, E.CODE_COLUMNS, 2 + i, ex)

    # local_id-Dropdowns auf den verknüpften Blättern
    for ws, n in ((flow_ws, flow_rows), (codes_ws, code_rows)):
        dv = DataValidation(type="list", formula1=id_range["element_ids"], allow_blank=True)
        dv.add(f"A2:A{n + 1}")
        ws.add_data_validation(dv)

    # Dropdowns der Datenelemente nachziehen (Codelisten existieren jetzt)
    for ci, (_k, _h, _m, listname) in enumerate(E.ELEMENT_COLUMNS, start=1):
        if listname and listname in ranges:
            suggestion = listname in E.SUGGESTION_LISTS
            d = DataValidation(type="list", formula1=ranges[listname], allow_blank=True,
                               showErrorMessage=not suggestion)
            if not suggestion:
                d.errorTitle, d.error, d.errorStyle = "Ungültige Eingabe", "Bitte aus der Liste wählen.", "stop"
            col = get_column_letter(ci)
            d.add(f"{col}2:{col}{elem_rows + 1}")
            elem_ws.add_data_validation(d)

    wb.move_sheet("Codelisten", offset=len(wb.sheetnames))  # ans Ende
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=str(E.ROOT / "templates" / "datenelement-erhebung.xlsx"))
    ap.add_argument("--rows", type=int, default=60, help="Echte Eingabezeilen im Blatt 'Datenelemente'")
    args = ap.parse_args()
    out = build(Path(args.out), args.rows)
    print(f"Wrote {out} (Datenelemente: 1 Beispiel + {args.rows} Zeilen; Datennutzung + Codes inkl. Beispiel)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
