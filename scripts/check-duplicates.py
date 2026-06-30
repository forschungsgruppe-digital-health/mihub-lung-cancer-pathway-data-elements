#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Prüft Datenelement-Kandidaten auf Dubletten/Ähnlichkeit gegen den Bestand.

Deterministischer Backbone für die Dubletten-Prüfung (Skill / Sub-Agent
`check-duplicate-data-element`). Vergleicht Kandidaten — aus neuen GitHub-Issues,
ausgefüllten Excel-Mappen (bzw. deren YAML-Entwürfen) oder neuen YAMLs — gegen alle
`elements/**/*.yaml` und untereinander. Signale: identischer `name`/`id`,
gemeinsames Coding (SNOMED/LOINC/ICD…), Wortähnlichkeit von `label_de`/`definition_de`,
gleiche `phase`.

Aufruf:
  python scripts/check-duplicates.py                         # Bestand gegen sich selbst (interne Nähe)
  python scripts/check-duplicates.py --candidates elements/_incoming/
  python scripts/check-duplicates.py --candidates neu.yaml
  python scripts/check-duplicates.py --from-issue "Rauchstatus / Tabakkonsum des Patienten"
  python scripts/check-duplicates.py --from-issue-file issue_body.md   # GitHub-Issue-Formular
  python scripts/check-duplicates.py --from-xlsx ausgefuellt.xlsx      # benötigt openpyxl

Exit-Code 1, wenn mind. ein Kandidat als DUBLETTE eingestuft wird (CI-tauglich).
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import yaml

ROOT = Path(__file__).resolve().parent.parent
ELEMENTS = ROOT / "elements"

_UML = str.maketrans({"ä": "a", "ö": "o", "ü": "u", "ß": "s"})
STOP = {
    "der", "die", "das", "und", "oder", "in", "im", "des", "den", "dem", "ein", "eine",
    "einer", "eines", "bei", "von", "zur", "zum", "auf", "mit", "fur", "für", "pro", "je",
    "ist", "wird", "werden", "nach", "vor", "the", "of", "and", "a", "an", "patient",
    "patienten", "patientin", "lungenkarzinom", "nsclc", "sclc",
}
DUP, NEAR, REL, DIST = "DUBLETTE", "NAH-DUBLETTE", "VERWANDT", "EIGENSTÄNDIG"


def tokens(text: str) -> set[str]:
    text = (text or "").lower().translate(_UML)
    return {t for t in re.split(r"[^a-z0-9]+", text) if len(t) > 2 and t not in STOP}


def _prefix_match(x: str, y: str) -> bool:
    """Toleriert deutsche Komposita/Plurale: Tabakkonsum~Tabakkonsumstatus, Zigarette~Zigaretten."""
    return x == y or (len(x) >= 4 and len(y) >= 4 and (x.startswith(y) or y.startswith(x)))


def soft_overlap(a: set[str], b: set[str]) -> float:
    """Jaccard mit Präfix-Toleranz (Stemming-Ersatz für Komposita/Plurale)."""
    if not a or not b:
        return 0.0
    inter = sum(1 for x in a if any(_prefix_match(x, y) for y in b))
    union = len(a) + len(b) - inter
    return inter / union if union else 0.0


def codings_of(doc: dict) -> set[str]:
    out = set()
    for c in ((doc.get("value_set") or {}).get("codings") or []):
        sys_, code = c.get("system"), c.get("code")
        if sys_ and code and sys_ != "local-no-standard-binding":
            out.add(f"{sys_}|{code}")
    return out


def descriptor(doc: dict, *, source: str = "") -> dict:
    return {
        "id": doc.get("id", ""),
        "name": doc.get("name", ""),
        "label_de": doc.get("label_de", ""),
        "definition_de": doc.get("definition_de", ""),
        "phase": doc.get("phase", ""),
        "codings": codings_of(doc),
        "label_tok": tokens(doc.get("label_de", "")),
        "def_tok": tokens(doc.get("definition_de", "")),
        "source": source,
    }


def load_corpus() -> list[dict]:
    corpus = []
    for fp in sorted(ELEMENTS.glob("*/*.yaml")):
        if "_incoming" in fp.parts:
            continue
        doc = yaml.safe_load(fp.read_text(encoding="utf-8")) or {}
        d = descriptor(doc, source=str(fp.relative_to(ROOT)))
        corpus.append(d)
    return corpus


def score(cand: dict, ref: dict) -> tuple[float, str, list[str]]:
    signals: list[str] = []
    if cand.get("name") and cand["name"] == ref["name"]:
        return 1.0, DUP, [f"identischer name '{cand['name']}'"]
    if cand.get("id") and cand["id"] == ref["id"]:
        return 1.0, DUP, [f"identische id '{cand['id']}'"]
    shared = cand["codings"] & ref["codings"]
    label_j = soft_overlap(cand["label_tok"], ref["label_tok"])
    def_j = soft_overlap(cand["def_tok"], ref["def_tok"])
    phase_match = bool(cand.get("phase")) and cand["phase"] == ref["phase"]
    s = 0.45 * label_j + 0.45 * def_j + (0.10 if phase_match else 0.0)
    lt, rt = cand["label_tok"], ref["label_tok"]
    if lt and rt and (lt <= rt or rt <= lt):  # eine Bezeichnung ist vollständig in der anderen enthalten
        s = max(s, 0.62)
        signals.append("Bezeichnung ist in der anderen vollständig enthalten")
    if shared:
        s = max(s, 0.8)
        signals.append("gemeinsame Codings: " + ", ".join(sorted(shared)))
    if label_j:
        signals.append(f"label-Ähnlichkeit {label_j:.0%}")
    if def_j:
        signals.append(f"Definition-Ähnlichkeit {def_j:.0%}")
    if phase_match:
        signals.append(f"gleiche Phase '{cand['phase']}'")
    verdict = DUP if s >= 0.85 else NEAR if s >= 0.6 else REL if s >= 0.4 else DIST
    return s, verdict, signals


def report(cand: dict, corpus: list[dict], top: int) -> str:
    ranked = sorted(((score(cand, r), r) for r in corpus), key=lambda x: x[0][0], reverse=True)
    best_verdict = ranked[0][0][1] if ranked else DIST
    title = cand.get("label_de") or cand.get("name") or cand.get("source") or "(Kandidat)"
    lines = [f"### Kandidat: {title}  →  **{best_verdict}**"]
    if cand.get("source"):
        lines.append(f"_Quelle: {cand['source']}_")
    shown = [(sc, r) for sc, r in ranked if sc[0] >= 0.4][:top]
    if not shown:
        lines.append("Keine ähnlichen Bestands-Elemente (≥40 %) gefunden — voraussichtlich eigenständig.")
    else:
        lines.append("| Score | Einstufung | Bestands-Element | Signale |")
        lines.append("| --- | --- | --- | --- |")
        for (sc, verdict, signals), r in shown:
            lines.append(f"| {sc:.2f} | {verdict} | `{r['name']}` ({r['phase']}) | {'; '.join(signals) or '—'} |")
    return "\n".join(lines)


def parse_issue_form(text: str) -> dict:
    """Extrahiert Bezeichnung + Definition aus dem GitHub-Issue-Formular (### Header)."""
    label, definition = "", ""
    blocks = re.split(r"^###\s+", text, flags=re.M)
    for b in blocks:
        head, _, body = b.partition("\n")
        h, body = head.strip().lower(), body.strip()
        if "bezeichnung" in h:
            label = body.splitlines()[0].strip() if body else ""
        elif "definition" in h:
            definition = body
    if not (label or definition):  # Freitext-Fallback
        label = text.strip().splitlines()[0][:120] if text.strip() else ""
        definition = text.strip()
    return {"label_de": label, "definition_de": definition}


def candidates_from_args(args) -> list[dict]:
    cands: list[dict] = []
    if args.from_issue:
        d = {"label_de": args.from_issue.strip(), "definition_de": args.from_issue.strip()}
        cands.append(descriptor(d, source="--from-issue"))
    if args.from_issue_file:
        d = parse_issue_form(Path(args.from_issue_file).read_text(encoding="utf-8"))
        cands.append(descriptor(d, source=args.from_issue_file))
    if args.from_xlsx:
        from openpyxl import load_workbook  # nur hier benötigt
        import _elicitation as _E
        wb = load_workbook(args.from_xlsx, data_only=True)
        ws = wb["Datenelemente"]
        keys = [c[0] for c in _E.ELEMENT_COLUMNS]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {keys[i]: row[i] for i in range(min(len(keys), len(row)))}
            if rec.get("label_de"):
                cands.append(descriptor(
                    {"label_de": rec.get("label_de", ""), "definition_de": rec.get("definition_de", "")},
                    source=f"xlsx:{rec.get('local_id') or rec.get('label_de')}"))
    for path in args.candidates or []:
        p = Path(path)
        files = sorted(p.glob("*.yaml")) if p.is_dir() else [p]
        for fp in files:
            doc = yaml.safe_load(fp.read_text(encoding="utf-8")) or {}
            cands.append(descriptor(doc, source=str(fp)))
    return cands


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--candidates", nargs="*", help="YAML-Datei(en) oder Verzeichnis(se) mit Kandidaten")
    ap.add_argument("--from-issue", help="Freitext (z. B. Issue-Titel/-Beschreibung)")
    ap.add_argument("--from-issue-file", help="Datei mit GitHub-Issue-Body (Formular)")
    ap.add_argument("--from-xlsx", help="Ausgefüllte Excel-Erhebungs-Mappe (benötigt openpyxl)")
    ap.add_argument("--top", type=int, default=5, help="max. Treffer je Kandidat")
    args = ap.parse_args()

    corpus = load_corpus()
    cands = candidates_from_args(args)
    self_mode = not cands
    if self_mode:  # Bestand gegen sich selbst (ohne Selbstvergleich)
        cands = corpus

    print(f"# Dubletten-Bericht — {len(cands)} Kandidat(en) gegen {len(corpus)} Bestands-Elemente\n")
    any_dup = False
    for cand in cands:
        pool = [r for r in corpus if r is not cand] if self_mode else corpus
        ranked = sorted(((score(cand, r), r) for r in pool), key=lambda x: x[0][0], reverse=True)
        if ranked and ranked[0][0][1] == DUP:
            any_dup = True
        print(report(cand, pool, args.top) + "\n")
    print("---\n> Heuristik (Wort-Jaccard + Coding-Overlap). Die Letztentscheidung über Dublette/"
          "Zusammenführung trifft ein Mensch (siehe Skill `check-duplicate-data-element`).")
    return 1 if any_dup else 0


if __name__ == "__main__":
    raise SystemExit(main())
